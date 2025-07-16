import pandas as pd
import openpyxl as px
from datetime import datetime
import os

#Excelファイルを取り出す
Filepath = os.path.abspath('/Users/sungyongan/Desktop/sample1.xlsx') 

#Excelファイルをpandasで読み込む
df = pd.read_excel(Filepath)  

#年-月-日から年-月に表示変更
df['月日'] = pd.to_datetime(df['月日']).dt.strftime("%Y-%m")

#それぞれの列で重複削除したリストを作成
D = list(df['月日'].unique()) 
P = list(df['製品'].unique())
F = list(df['利益'].unique())
Pl = list(df['支店'].unique())

#現在の日付を取得
now = datetime.now()
hiduke = now.strftime('%Y-%m-%d')

#製品ごとに分けたデータのシートを生成
for products in P: 
    filtered = df[df['製品'] == f'{products}']
    sales = pd.pivot_table(df, index=filtered['月日'], columns='支店', values='利益', aggfunc='sum', fill_value=0)

    #ファイル名（{hiduke}には現在の日付{products}には商品名）
    filepath = f'/Users/sungyongan/Desktop/{hiduke}_{products}.xlsx' 
    sales.to_excel(filepath, sheet_name='月別総合利益', startrow=3)
    wb = px.load_workbook(filepath)

    #シート名の取得
    ws = wb['月別総合利益'] 

    #A1のセルに「{products}_売上」と記入
    ws.cell(row=1, column=1).value = f'{products}_売上' 

    #フォント設定
    ws.cell(row=1, column=1).font = px.styles.Font(size=12, bold=True) 

    #A2のセルに「月次売り上げ」と記入
    ws.cell(row=2, column=1).value = '月次売り上げ' 

    #フォント設定
    ws.cell(row=2, column=1).font = px.styles.Font(size=12, bold=True)

    #折れ線グラフの生成
    chart = px.chart.LineChart() 
    data = px.chart.Reference(ws, min_col=2, max_col=len(Pl)+1, min_row=4, max_row=len(D)+4)
    categories = px.chart.Reference(ws, min_col=1, max_col=1, min_row=5, max_row=len(D)+4)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    #グラフ表示の大きさ、グラフの題名、グラフの単位を記入する
    chart.style = 14 
    chart.title = '製品売上'
    chart.y_axis.title = '円'
    chart.x_axis.title = '年月'
    chart.height = 9
    chart.width = 16
    #グラフの生成

    ws.add_chart(chart, "G2")

    #各ファイルの保存
    wb.save(filepath)
