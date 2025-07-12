from openpyxl import load_workbook


def check_num_target_cells(sheet, check_row_num, check_column_num):

    num_target_cells = 0
    mail_address_list = []
    user_name_list = []

    while sheet.cell(row=check_row_num, column=check_column_num).value != None:
        mail_address_list.append(sheet.cell(row=check_row_num, column=check_column_num).value)
        user_name_list.append(sheet.cell(row=check_row_num, column=check_column_num + 1).value)
        check_row_num += 1
        num_target_cells += 1

    return num_target_cells, mail_address_list, user_name_list


def check_account(mail_address, user_name):

    excel_path = "sample.xlsx"
    wb = load_workbook(filename=excel_path, read_only=True)
    response_content_flag = False
    response_content = ""

    sheet = wb.worksheets[1]

    check_row_num = 3
    check_column_num = 1
    num_target_cells, mail_address_list, user_name_list = check_num_target_cells(sheet, check_row_num, check_column_num)

    if mail_address in mail_address_list and user_name in user_name_list:
        for i in range(num_target_cells):
            print(mail_address_list[i], user_name_list[i])
            # 入力したメールアドレスと対象セルのメールアドレスが一致する場合
            if mail_address == mail_address_list[i]:
                if user_name == user_name_list[i]:
                    return response_content_flag, response_content
                # ユーザー名が不一致の場合
                else:
                    response_content_flag = True
                    response_content = "入力した情報に誤りがあります。管理者に確認してください。"
                    return response_content_flag, response_content
    else:
        response_content_flag = True
        response_content = "入力したユーザー名、もしくはメールアドレスに誤りがあります。"
        return response_content_flag, response_content

if __name__ == "__main__":
    response_content_flag, response_content = check_account("1@1", "大谷")

    if response_content_flag == True:
        print(response_content)

    else:
        print("成功", response_content)
