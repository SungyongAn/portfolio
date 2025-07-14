from openpyxl import load_workbook


# ユーザー名から情報を記載する行の確認
def check_target_num_row(sheet, user_name):
    # 作業日報の作業Aの名前登録の最上段（固定）、他作業も同行での入力となるため作業Aをから行を特定する。
    check_row_num = 3
    check_column_num = 1

    while sheet.cell(row=check_row_num, column=check_column_num).value != user_name:
        check_row_num += 1

    return check_row_num


# Excelファイルへの作業内容の書込み
def write_to_excel(today_date, user_name, work_type, time_worked):

    excel_path = "sample.xlsx" # 絶対path

    wb = load_workbook(filename=excel_path)

    # Excelファイルへ正常に入力できたか確認用のフラグ
    response_content_flag = False
    # 正常に入力されているかどうかの確認用リスト、入力前のセルの中身を記録する。
    for_check_list_before = ""

    sheet = wb.worksheets[0]

    work_type_dict = {"A": 1, "B": 9, "C": 21}

    # ユーザー名から情報を記載する行の確認
    target_num_row = check_target_num_row(sheet, user_name)

    # 日付を入力するセル列
    date_cell_column = work_type_dict[work_type[0]] + 1
    # 入力前の日付を保存
    for_check_list_before = sheet.cell(row=target_num_row, column=date_cell_column).value
    # 日付の入力
    sheet.cell(row=target_num_row, column=date_cell_column, value=today_date)

    # 作業回数を加算するセル列
    num_of_operations_cell_column = work_type_dict[work_type[0]] + 2
    # 入力前の作業回数を保存
    for_check_list_before = sheet.cell(row=target_num_row, column=num_of_operations_cell_column).value
    # 作業回数の加算
    if sheet.cell(row=target_num_row, column=num_of_operations_cell_column).value == None:
        sheet.cell(row=target_num_row, column=num_of_operations_cell_column, value=1)
    else:
        num_of_tasks_this_time = sheet.cell(row=target_num_row, column=num_of_operations_cell_column).value + 1
        sheet.cell(row=target_num_row, column=num_of_operations_cell_column, value=num_of_tasks_this_time)

    # 作業回数の更新が正常に行われたか確認
    if for_check_list_before + 1 != sheet.cell(row=target_num_row, column=num_of_operations_cell_column).value:
        response_content_flag == True
        response_content = "エラーが発生しました。"
        return response_content

    # 作業時間を記入するセル列
    time_worked_cell_column = work_type_dict[work_type[0]] + 2 + int(work_type[1])
    # 入力前の作業時間を保存
    for_check_list_before = sheet.cell(row=target_num_row, column=time_worked_cell_column).value
    # 作業時間の記入
    if sheet.cell(row=target_num_row, column=time_worked_cell_column).value == None:
        sheet.cell(row=target_num_row, column=time_worked_cell_column, value=time_worked)
    else:
        total_time_worked = sheet.cell(row=target_num_row, column=time_worked_cell_column).value + time_worked
        sheet.cell(row=target_num_row, column=time_worked_cell_column, value=total_time_worked)
    
    # 作業時間の入力が正常に行われたか確認
    if for_check_list_before == sheet.cell(row=target_num_row, column=time_worked_cell_column, value=total_time_worked):
        response_content_flag == True
        response_content = "エラーが発生しました。"
        return response_content

    wb.save("sample.xlsx") # 絶対path

    response_content = "お疲れ様でした。"

    return response_content


if __name__ == "__main__":
    today_date = "2025/07/14"
    user_name = "桑田"
    work_type = "A1"
    time_worked = 60

    response_content = write_to_excel(today_date, user_name, work_type, time_worked)

    print(response_content)
