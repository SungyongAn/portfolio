from openpyxl import load_workbook


# メールアドレスとユーザー名の登録確認
def check_account(mail_address, user_name, work_flag):

    # 対象のExcelファイルを開く（絶対PATH）
    excel_path = "/Users/sungyongan/workspace/portfolio/excel_automation/backend/routes/sample.xlsx" 
    wb = load_workbook(filename=excel_path, read_only=True)
    # アカウント管理用のシートを指定
    sheet = wb.worksheets[1]

    # アカウント情報の記載の最上段のセル位置（固定）
    check_row_num = 3
    check_column_num = 1

    # ループ処理時に確認する範囲の初期値
    num_target_cells = 0
    # 登録済みメールアドレスの一覧作成用
    mail_address_list = []
    # 登録済みユーザー名の一覧作成用
    user_name_list = []

    while sheet.cell(row=check_row_num, column=check_column_num).value != None:
        mail_address_list.append(sheet.cell(row=check_row_num, column=check_column_num).value)
        user_name_list.append(sheet.cell(row=check_row_num, column=check_column_num + 1).value)
        check_row_num += 1
        num_target_cells += 1

    if mail_address in mail_address_list and user_name in user_name_list:
        for i in range(num_target_cells):
            # 入力したメールアドレスと対象セルのメールアドレスが一致する場合
            if mail_address == mail_address_list[i]:
                if user_name == user_name_list[i]:
                    response_content = "以下の項目を確認、入力の上、最後に送信ボタンを押してください。"
                    return response_content, mail_address, user_name
                # ユーザー名が不一致の場合
                else:
                    response_content = "入力した情報に誤りがあります。管理者に確認してください。"
                    return response_content, mail_address, user_name
    else:
        response_content = "入力したユーザー名、もしくはメールアドレスに誤りがあります。"
        return response_content, mail_address, user_name
