import openpyxl as px
from datetime import datetime, timedelta 
from openpyxl.worksheet.table import Table, TableStyleInfo
import random


def create_table_in_excel(num_add_sheets, start_date_str, end_date_str, filename):

    start_date = datetime.strptime(start_date_str, "%Y-%m-%d")
    end_date = datetime.strptime(end_date_str, "%Y-%m-%d")

    wb = px.Workbook()
    ws = wb.active
    num_sheets = num_add_sheets + 1

    # シートの追加
    if num_add_sheets != 0:
        for i in range(num_add_sheets):
            num = i + 1
            wb.create_sheet(title=f"sheet_{num}")

        # 初期配置のsheetの名称を他に合わせて変更（変更手順の確認を兼ねて）
        wb.worksheets[0].title = "sheet_0"

    for i in range(num_sheets):
        ws = wb.worksheets[i]
        ws['A1'] = '月日'
        ws['B1'] = '製品'
        ws['C1'] = '支店'
        ws['D1'] = '利益'

    # sheet_0のみ日付の記入
    current_date = start_date
    target_sheet = 0
    row_num = 2

    while current_date <= end_date:
        ws = wb.worksheets[target_sheet]
        ws.cell(row=row_num, column=column_num, value=current_date.strftime("%Y-%m-%d"))
        month_zero = current_date.month
        current_date += timedelta(days=1)
        month = current_date.month
        row_num += 1
        # 月を跨いだ時に列をずらして一番上から入力する。
        if month != month_zero:
            column_num += 1
            row_num = 1
        if current_date == end_date:
            num_target_cells = row_num

    for k in range(2, num_target_cells + 1):
        ws = wb.worksheets[0]
        ws['B'+ str(k)] = random.choice(['製品1', '製品2'])
        ws['C'+ str(k)] = random.choice(['支店1', '支店2'])
        ws['D'+ str(k)] = random.randint(10000, 100000)

    ws.column_dimensions['A'].width = 13

    # テーブルを生成
    table = Table(displayName='Table1', ref=f'A1:D{num_target_cells}')
    table.tableStyleInfo = TableStyleInfo(name='TableStyleMedium12', showRowStripes=True)
    ws.add_table(table)

    # 直接保存
    wb.save(filename)


if __name__ == "__main__":
    num_add_sheets = 2
    start_date_str = "2025-01-01"
    end_date_str = "2025-03-31"
    filename = "sample2_テーブル.xlsx"
    create_table_in_excel(num_add_sheets, start_date_str, end_date_str, filename)
